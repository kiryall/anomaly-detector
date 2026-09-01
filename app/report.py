from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.detection import DetectionResult
from app.services.state import Statistics


def _safe_class_name(class_name: str) -> str:
    """Возвращает безопасное имя для директории."""
    name = class_name.strip().lower()
    name = re.sub(r"[^\w\-]", "_", name)
    name = re.sub(r"_+", "_", name)
    name = name.lstrip("_").rstrip("_")
    if not name:
        name = "unknown"
    if ".." in name or name.startswith("/"):
        name = "unknown"
    return name


class ReportService:
    """Формирует Excel-отчёт по результатам детекции."""

    # ==========================================================
    # Публичный интерфейс
    # ==========================================================

    def save_excel(
        self,
        results: tuple[DetectionResult, ...],
        statistics: Statistics,
        output_path: Path,
        model_name: str,
        source_folder: Path,
        confidence_threshold: float,
    ) -> Path:
        """Создаёт Excel-отчёт и сохраняет его.

        Args:
            results: Результаты детекции текущего запуска.
            statistics: Статистика обработки.
            output_path: Путь сохранения отчёта.
            model_name: Имя использованной модели.
            source_folder: Папка с исходными изображениями.
            confidence_threshold: Порог уверенности.

        Returns:
            Путь к сохранённому файлу.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        self._populate_summary(
            wb,
            results,
            statistics,
            model_name,
            source_folder,
            confidence_threshold,
        )

        self._populate_detections(wb, results)

        wb.save(str(output_path))

        return output_path

    # ==========================================================
    # Лист "Сводка"
    # ==========================================================

    def _populate_summary(
        self,
        wb: Workbook,
        results: tuple[DetectionResult, ...],
        statistics: Statistics,
        model_name: str,
        source_folder: Path,
        confidence_threshold: float,
    ) -> None:
        """Заполняет лист сводной информацией."""
        ws = wb.active
        ws.title = "Сводка"

        # Заголовок
        ws.append(["Отчёт детекции"])

        ws.append([])

        # Информация о запуске
        ws.append(["Модель", model_name])
        ws.append(["Папка с исходными изображениями", str(source_folder)])
        ws.append(["Confidence threshold", f"{confidence_threshold:.2f}"])
        ws.append([
            "Дата и время формирования отчёта",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ])

        ws.append([])

        # Статистика
        ws.append(["Статистика обработки"])
        ws.append([])

        total_images = len(results)
        anomaly_images = sum(1 for r in results if r.has_anomaly)
        normal_images = total_images - anomaly_images
        total_detections = sum(len(r.detections) for r in results)
        multi_class_images = sum(
            1 for r in results
            if len({d.class_id for d in r.detections}) >= 2
        )
        total_errors = statistics.errors
        total_time = sum(r.inference_time for r in results)
        avg_time = total_time / total_images if total_images > 0 else 0.0

        ws.append(["Всего обработано", total_images])
        ws.append(["Фото с дефектами", anomaly_images])
        ws.append(["Фото без дефектов", normal_images])
        ws.append(["Всего обнаружено дефектов", total_detections])
        ws.append(["Фото с несколькими классами", multi_class_images])
        ws.append(["Ошибок", total_errors])
        ws.append(["Общее время детекции, сек", f"{total_time:.3f}"])
        ws.append(["Среднее время детекции, сек", f"{avg_time:.3f}"])

        # Форматирование
        self._format_summary_sheet(ws)

    def _format_summary_sheet(self, ws: object) -> None:
        """Форматирует лист сводки."""
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)

        # Заголовок
        if ws.max_row >= 1:
            ws.cell(1, 1).font = header_font

        # Секция "Статистика обработки"
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, 1)
            if cell.value == "Статистика обработки":
                cell.font = section_font

    # ==========================================================
    # Лист "Детекции"
    # ==========================================================

    def _populate_detections(
        self,
        wb: Workbook,
        results: tuple[DetectionResult, ...],
    ) -> None:
        """Заполняет лист детекций."""
        ws = wb.create_sheet("Детекции")

        # Заголовки колонок
        headers = [
            "Изображение",
            "Исходный путь",
            "Класс",
            "Class ID",
            "Confidence",
            "X1",
            "Y1",
            "X2",
            "Y2",
            "Папка результата",
            "Путь результата",
            "Время детекции, сек",
        ]
        ws.append(headers)

        # Данные
        for result in results:
            image_name = result.image_path.name
            image_path_str = str(result.image_path)

            if result.detections:
                # Для каждой детекции — отдельная строка
                for detection in result.detections:
                    # Определяем папку результата для данного класса
                    output_path = self._find_output_path_for_class(
                        result.output_paths,
                        detection.class_name,
                    )

                    row = [
                        image_name,
                        image_path_str,
                        detection.class_name,
                        detection.class_id,
                        detection.confidence,
                        detection.bbox.x1,
                        detection.bbox.y1,
                        detection.bbox.x2,
                        detection.bbox.y2,
                        output_path.parent.name if output_path else "",
                        str(output_path) if output_path else "",
                        result.inference_time,
                    ]
                    ws.append(row)
            else:
                # no_defect — одна строка без детекций
                output_path = (
                    result.output_paths[0] if result.output_paths else None
                )
                row = [
                    image_name,
                    image_path_str,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "no_defect",
                    str(output_path) if output_path else "",
                    result.inference_time,
                ]
                ws.append(row)

        # Форматирование
        self._format_detections_sheet(ws)

    @staticmethod
    def _find_output_path_for_class(
        output_paths: tuple[Path, ...],
        class_name: str,
    ) -> Path | None:
        """Находит путь результата для заданного класса."""
        safe_name = _safe_class_name(class_name)
        for path in output_paths:
            if path.parent.name == safe_name:
                return path
        return None

    def _format_detections_sheet(self, ws: object) -> None:
        """Форматирует лист детекций."""
        from openpyxl.worksheet.worksheet import Worksheet

        if not isinstance(ws, Worksheet):
            return

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        number_format_confidence = "0.000"
        number_format_time = "0.000"
        number_format_coord = "0.00"

        # Форматируем заголовки
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Числовой формат для Confidence, X1-Y2, Время
        for row_idx in range(2, ws.max_row + 1):
            # Confidence (E)
            cell_e = ws.cell(row_idx, 5)
            cell_e.number_format = number_format_confidence

            # X1 (F)
            cell_f = ws.cell(row_idx, 6)
            cell_f.number_format = number_format_coord

            # Y1 (G)
            cell_g = ws.cell(row_idx, 7)
            cell_g.number_format = number_format_coord

            # X2 (H)
            cell_h = ws.cell(row_idx, 8)
            cell_h.number_format = number_format_coord

            # Y2 (I)
            cell_i = ws.cell(row_idx, 9)
            cell_i.number_format = number_format_coord

            # Время детекции (L)
            cell_l = ws.cell(row_idx, 12)
            cell_l.number_format = number_format_time

        # Freeze panes
        ws.freeze_panes = "A2"

        # AutoFilter
        ws.auto_filter.ref = f"A1:{self._column_letter(ws.max_column)}{ws.max_row}"

        # Auto-width (приблизительная)
        column_widths = {
            "A": 15,   # Изображение
            "B": 50,   # Исходный путь
            "C": 15,   # Класс
            "D": 10,   # Class ID
            "E": 12,   # Confidence
            "F": 10,   # X1
            "G": 10,   # Y1
            "H": 10,   # X2
            "I": 10,   # Y2
            "J": 15,   # Папка результата
            "K": 50,   # Путь результата
            "L": 18,   # Время детекции
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    @staticmethod
    def _column_letter(column_idx: int) -> str:
        """Преобразует номер колонки в буквенное обозначение."""
        result = ""
        while column_idx > 0:
            column_idx, remainder = divmod(column_idx - 1, 26)
            result = chr(ord("A") + remainder) + result
        return result
